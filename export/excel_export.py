import pandas as pd
from io import BytesIO
from flask import send_file, request
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from export.prepare_dataframe import prepare_export_dataframe
from export.export_translator import translate_export

def create_excel_report(df):
    df = prepare_export_dataframe(df)

    lang = request.cookies.get("language", "bg")
    df = translate_export(df, lang)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="CPU Recommendations",index=False)
        worksheet = writer.sheets["CPU Recommendations"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        header_fill = PatternFill(fill_type="solid",fgColor="4472C4")

        for cell in worksheet[1]:
            cell.font = Font(bold=True,color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center",vertical="center")

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 3, 35)

        headers = [cell.value for cell in worksheet[1]]

        for row in range(2, worksheet.max_row + 1):
            for col_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row, column=col_index)
                if "Price" in str(header) or "Цена" in str(header):
                    cell.number_format = '0.00 €'
                elif "TOPSIS" in str(header):
                    cell.number_format = "0.00"

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"CPU_Recommendations_{datetime.now():%Y-%m-%d_%H-%M}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )